"""
Módulo para atualizar planilhas de orçamento com dados correlacionados
Utiliza openpyxl para preservar fórmulas e estrutura original
"""

import os
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import List, Dict
import logging
import sys

# Configurar logging
Path("logs").mkdir(exist_ok=True)
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

console_handler = logging.StreamHandler(sys.stdout)
file_handler = logging.FileHandler(f"logs/automacao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

formatter = logging.Formatter(
    '%(asctime)s | %(levelname)-8s | %(funcName)s():%(lineno)d | %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

console_handler.setFormatter(formatter)
file_handler.setFormatter(formatter)
logger.addHandler(console_handler)
logger.addHandler(file_handler)


class AtualizadorPlanilha:
    """Atualiza planilha de orçamento com dados correlacionados preservando fórmulas"""
    
    def __init__(self, caminho_planilha: str, aba: str, parametros: Dict):
        """
        Inicializa o atualizador.
        
        Args:
            caminho_planilha: Caminho da planilha original
            aba: Nome da aba a atualizar
            parametros: Dict com colunas (descrição, material, mao_de_obra, unidade_medida)
        """
        self.caminho_planilha = caminho_planilha
        self.aba = aba
        self.parametros = parametros
        self.workbook = None
        self.worksheet = None
    
    def _converter_coluna_para_indice(self, coluna: str) -> int:
        """Converte letra de coluna (A, B, C...) para índice (1, 2, 3...)"""
        coluna = coluna.upper()
        indice = 0
        for char in coluna:
            indice = indice * 26 + (ord(char) - ord('A') + 1)
        return indice
    
    def _carregar_planilha(self):
        """Carrega a planilha original"""
        try:
            self.workbook = load_workbook(self.caminho_planilha)
            self.worksheet = self.workbook[self.aba]
            logger.debug(f"Planilha carregada: {self.aba}")
        except Exception as e:
            raise Exception(f"Erro ao carregar planilha: {e}")
    
    def _salvar_planilha_atualizada(self, caminho_destino: str):
        """Salva a planilha atualizada"""
        try:
            self.workbook.save(caminho_destino)
            logger.debug(f"Planilha salva em: {caminho_destino}")
            return True
        except Exception as e:
            raise Exception(f"Erro ao salvar planilha: {e}")
    
    def _gerar_caminho_com_timestamp(self, caminho_original: str) -> str:
        """Gera caminho com timestamp para evitar sobrescrita"""
        base = Path(caminho_original)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        novo_nome = f"{base.stem}_PREENCHIDA_{timestamp}{base.suffix}"
        return str(base.parent / novo_nome)
    
    def atualizar_com_selecoes(self, selecoes: List[Dict], caminho_destino: str = None) -> str:
        """
        Atualiza a planilha com os itens selecionados.
        
        Args:
            selecoes: Lista de dicts com {
                "item": nome do item,
                "numero_linha": número da linha na planilha,
                "referencia": descrição da referência,
                "valor_material": valor do material,
                "valor_mao_de_obra": valor da mão de obra,
                "unidade": unidade de medida
            }
            caminho_destino: Onde salvar (se None, usa timestamp automático)
        
        Returns:
            Caminho do arquivo salvo
        """
        logger.debug(f"Atualizando planilha com {len(selecoes)} itens")
        
        # Carregar planilha
        self._carregar_planilha()
        
        # Converter colunas para indices
        col_material = self._converter_coluna_para_indice(self.parametros["coluna_material"])
        col_mao_de_obra = self._converter_coluna_para_indice(self.parametros["coluna_mao_de_obra"])
        col_unidade = self._converter_coluna_para_indice(self.parametros["coluna_unidade_medida"])
        
        logger.debug(f"Colunas: Material={col_material}, Mao de Obra={col_mao_de_obra}, Unidade={col_unidade}")
        
        # Atualizar células
        atualizacoes = 0
        for selecao in selecoes:
            numero_linha = selecao.get("numero_linha")
            
            if not numero_linha:
                logger.warning(f"Numero de linha nao fornecido para {selecao['item']}")
                continue
            
            try:
                # Preencher material
                celula_material = self.worksheet.cell(row=numero_linha, column=col_material)
                celula_material.value = float(selecao["valor_material"])
                logger.debug(f"Linha {numero_linha}, Col {col_material}: {selecao['valor_material']}")
                
                # Preencher mao de obra
                celula_mao_de_obra = self.worksheet.cell(row=numero_linha, column=col_mao_de_obra)
                celula_mao_de_obra.value = float(selecao["valor_mao_de_obra"])
                logger.debug(f"Linha {numero_linha}, Col {col_mao_de_obra}: {selecao['valor_mao_de_obra']}")
                
                # Preencher unidade
                celula_unidade = self.worksheet.cell(row=numero_linha, column=col_unidade)
                celula_unidade.value = selecao["unidade"]
                logger.debug(f"Linha {numero_linha}, Col {col_unidade}: {selecao['unidade']}")
                
                atualizacoes += 1
            except Exception as e:
                logger.error(f"Erro ao atualizar linha {numero_linha}: {e}")
        
        logger.debug(f"{atualizacoes} linhas atualizadas com sucesso")
        
        # Definir caminho de salvamento
        if caminho_destino is None:
            caminho_destino = self._gerar_caminho_com_timestamp(self.caminho_planilha)
        
        # Salvar
        self._salvar_planilha_atualizada(caminho_destino)
        
        return caminho_destino
